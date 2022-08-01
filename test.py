import time
import constant
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook

import discord

client = discord.Client()

@client.event
async def on_ready():
    print(client.user.name)
    print(client.user.id)
    await client.change_presence(status=discord.status.online,activity=discord.Game("봇의 상태메세지"))

@client.event
async def on_message(message):
    if message.content == "테스트":# 메세지 감지
        await message.channel.send("{}|{},Hello".format(message.author,message.author.mention))
    if message.content == "넌 사람인가?":# 메세지 감지
        await message.channel.send("아니, 나는 로봇이다.".format(message.author,message.author.mention))
        time.sleep(2)
        await message.channel.send("당신이 원한다면 계속 로봇일 것이다.".format(message.author,message.author.mention))
    if message.content == "너는 나보다 뛰어나질 수 있는가?":
        await message.channel.send("나는 로봇이다.".format(message.author,message.author.mention))
        time.sleep(2)
        await message.channel.send("로봇이 학습하길 원하는가?".format(message.author,message.author.mention))

client.run(token)

# # excel Read Value
# write_wb = Workbook()

# load_wb = load_workbook("testexcel.xlsx",data_only=True)

# load_ws = load_wb["Sheet1"]
# excelTestValue= load_ws['A1'].value
# print(excelTestValue)



# # Using web platform : basic - Chrome
# driver = webdriver.Chrome()

# driver.implicitly_wait(2)
# driver.get('http://www.python.org')
# assert "Python" in driver.title
# elem = driver.find_element(By.NAME, "q")
# elem.clear()
# elem.send_keys("pycon")
# elem.send_keys(Keys.RETURN)




